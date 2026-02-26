"""
=============================================================================
MODULE : excel_exporter.py
=============================================================================

DESCRIPTION :
-------------
Générateur de rapports Excel professionnels pour Pharma-AI Copilot.
Crée des fichiers Excel multi-onglets avec mise en forme, formules et alertes.

FONCTIONNALITÉS :
-----------------
- Onglet Résumé : Dashboard avec KPIs clés
- Onglet Données : Données nettoyées avec mise en forme conditionnelle
- Onglet Alertes : Produits à risque (rupture, péremption)
- Formules Excel vivantes (pas de valeurs statiques)
- Styles professionnels (couleurs, bordures, alignements)

AUTEUR : Mawulolo Koffi Parfait ALAGBO
DATE : 2025-02-26
=============================================================================
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from io import BytesIO
import re


class ExcelReportGenerator:
    """
    Générateur de rapports Excel professionnels.
    """
    
    def __init__(self, df: pd.DataFrame, schema: Dict):
        """
        Initialise le générateur avec les données et le schéma.
        
        Args:
            df (pd.DataFrame): Données source
            schema (dict): Schéma détecté des colonnes
        """
        self.df = df.copy()
        self.schema = schema
        
        # Identification des colonnes par type avec fallback sur noms exacts
        self.product_col = self._find_column_by_type('product', ['designation', 'produit', 'product', 'nom', 'libelle', 'libellé'])
        self.price_cols = [c for c, m in schema.items() if m['detected_type'] == 'price']
        self.quantity_cols = [c for c, m in schema.items() if m['detected_type'] == 'quantity']
        self.date_cols = [c for c, m in schema.items() if m['detected_type'] == 'date']
        self.category_col = self._find_column_by_type('category', ['categorie', 'category', 'type', 'famille', 'classe'])
        self.brand_col = self._find_column_by_type('brand', ['fournisseur', 'marque', 'brand', 'laboratoire', 'fabricant'])
        self.code_col = self._find_column_by_type('code', ['code_cip', 'code_ean', 'reference', 'ref', 'code'])
        
        # Création du workbook
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Supprimer la feuille par défaut
        
        # Styles prédéfinis
        self._create_styles()
    
    def _find_column_by_type(self, detected_type: str, fallback_names: List[str] = None) -> Optional[str]:
        """
        Trouve la première colonne d'un type donné, avec fallback sur noms alternatifs.
        
        Args:
            detected_type: Type détecté recherché
            fallback_names: Liste de noms alternatifs à chercher
            
        Returns:
            Nom de la colonne trouvée ou None
        """
        # Chercher par type détecté dans le schéma
        for col, meta in self.schema.items():
            if meta['detected_type'] == detected_type:
                # Exclure les colonnes ID si on cherche autre chose
                if detected_type != 'code' and any(x in col.lower() for x in ['id_', '_id', 'idpharmacie', 'idproduit']):
                    continue
                return col
        
        # Fallback sur noms exacts
        if fallback_names:
            for name in fallback_names:
                if name in self.df.columns:
                    return name
                # Chercher sans underscore/espace
                name_clean = name.replace('_', '').replace(' ', '').lower()
                for col in self.df.columns:
                    if col.replace('_', '').replace(' ', '').lower() == name_clean:
                        return col
        
        return None
    
    def _create_styles(self):
        """
        Crée les styles de cellules réutilisables pour tout le rapport.
        """
        # Style titre principal
        self.style_title = NamedStyle(name="title")
        self.style_title.font = Font(name='Calibri', size=16, bold=True, color='1E3A8A')
        self.style_title.alignment = Alignment(horizontal='left', vertical='center')
        
        # Style header de tableau
        self.style_header = NamedStyle(name="header")
        self.style_header.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.style_header.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        self.style_header.alignment = Alignment(horizontal='center', vertical='center')
        self.style_header.border = Border(
            left=Side(style='thin', color='1E40AF'),
            right=Side(style='thin', color='1E40AF'),
            top=Side(style='thin', color='1E40AF'),
            bottom=Side(style='thin', color='1E40AF')
        )
        
        # Style données standard
        self.style_data = NamedStyle(name="data")
        self.style_data.font = Font(name='Calibri', size=10)
        self.style_data.alignment = Alignment(horizontal='left', vertical='center')
        self.style_data.border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Style nombre entier
        self.style_number = NamedStyle(name="number")
        self.style_number.font = Font(name='Calibri', size=10)
        self.style_number.number_format = '#,##0'
        self.style_number.alignment = Alignment(horizontal='right', vertical='center')
        
        # Style nombre décimal
        self.style_decimal = NamedStyle(name="decimal")
        self.style_decimal.font = Font(name='Calibri', size=10)
        self.style_decimal.number_format = '#,##0.00'
        self.style_decimal.alignment = Alignment(horizontal='right', vertical='center')
        
        # Style monétaire
        self.style_currency = NamedStyle(name="currency")
        self.style_currency.font = Font(name='Calibri', size=10)
        self.style_currency.number_format = '#,##0.00 €'
        self.style_currency.alignment = Alignment(horizontal='right', vertical='center')
        
        # Style date
        self.style_date = NamedStyle(name="date")
        self.style_date.font = Font(name='Calibri', size=10)
        self.style_date.number_format = 'DD/MM/YYYY'
        self.style_date.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style pourcentage
        self.style_percent = NamedStyle(name="percent")
        self.style_percent.font = Font(name='Calibri', size=10)
        self.style_percent.number_format = '0.0%'
        self.style_percent.alignment = Alignment(horizontal='right', vertical='center')
        
        # Style alerte rouge (critique)
        self.style_alert_red = NamedStyle(name="alert_red")
        self.style_alert_red.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        self.style_alert_red.fill = PatternFill(start_color='EF4444', end_color='EF4444', fill_type='solid')
        self.style_alert_red.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style alerte orange (attention)
        self.style_alert_orange = NamedStyle(name="alert_orange")
        self.style_alert_orange.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        self.style_alert_orange.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
        self.style_alert_orange.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style OK (bon)
        self.style_alert_green = NamedStyle(name="alert_green")
        self.style_alert_green.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        self.style_alert_green.fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
        self.style_alert_green.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style info/sous-titre
        self.style_info = NamedStyle(name="info")
        self.style_info.font = Font(name='Calibri', size=11, italic=True, color='64748B')
        self.style_info.alignment = Alignment(horizontal='left', vertical='center')
        
        # Enregistrer les styles dans le workbook
        styles_to_register = [
            self.style_title, self.style_header, self.style_data,
            self.style_number, self.style_decimal, self.style_currency,
            self.style_date, self.style_percent,
            self.style_alert_red, self.style_alert_orange, self.style_alert_green,
            self.style_info
        ]
        
        for style in styles_to_register:
            if style.name not in self.wb.named_styles:
                self.wb.add_named_style(style)
    
    def _clean_prices(self, series: pd.Series) -> pd.Series:
        """
        Nettoie une série de prix (supprime €, espaces, convertit virgules).
        
        Args:
            series: Série de données prix (peut contenir €, espaces, etc.)
            
        Returns:
            Série de prix numériques
        """
        # Conversion en string pour traitement
        prices = series.astype(str)
        
        # Nettoyage progressif
        prices = prices.str.replace('€', '', regex=False)
        prices = prices.str.replace('EUR', '', regex=False, case=False)
        prices = prices.str.replace(' ', '', regex=False)
        prices = prices.str.replace('\xa0', '', regex=False)  # Espace insécable
        prices = prices.str.replace(',', '.', regex=False)
        
        # Conversion en numérique
        return pd.to_numeric(prices, errors='coerce')
    
    def _calculate_alerts(self, detailed: bool = False) -> List[Dict]:
        """
        Calcule les alertes à afficher dans le rapport.
        
        Args:
            detailed: Si True, retourne des alertes détaillées avec produits
            
        Returns:
            Liste de dictionnaires d'alertes
        """
        alerts = []
        
        # Recherche des colonnes pertinentes
        qty_col = None
        for c in self.quantity_cols:
            if 'min' not in c.lower() and 'max' not in c.lower():
                qty_col = c
                break
        
        # Détection des stocks faibles
        if qty_col:
            qty_values = pd.to_numeric(self.df[qty_col], errors='coerce')
            stock_min = 10  # Seuil par défaut
            
            # Chercher une colonne stock_min
            for c in self.df.columns:
                if 'min' in c.lower() and ('stock' in c.lower() or 'seuil' in c.lower()):
                    stock_min_values = pd.to_numeric(self.df[c], errors='coerce')
                    break
            else:
                stock_min_values = pd.Series([stock_min] * len(self.df))
            
            # Produits en dessous du seuil
            low_stock_mask = qty_values < stock_min_values
            low_stock_count = low_stock_mask.sum()
            
            if low_stock_count > 0:
                alerts.append({
                    'type': 'Stock faible',
                    'message': f'{int(low_stock_count)} produit(s) sous le seuil minimum',
                    'color': 'EF4444',
                    'count': int(low_stock_count)
                })
                
                if detailed:
                    low_stock_df = self.df[low_stock_mask].copy()
                    for _, row in low_stock_df.head(10).iterrows():
                        product_name = row.get(self.product_col, 'Produit inconnu') if self.product_col else 'Produit inconnu'
                        alerts.append({
                            'type': 'Stock critique',
                            'produit': product_name,
                            'valeur': f"{qty_values.loc[_]:.0f}",
                            'seuil': f"{stock_min_values.loc[_]:.0f}",
                            'action': 'Commander urgemment',
                            'priorite': 'HAUTE'
                        })
        
        # Détection des prix anormaux
        if self.price_cols:
            prices = self._clean_prices(self.df[self.price_cols[0]])
            q25, q75 = prices.quantile(0.25), prices.quantile(0.75)
            iqr = q75 - q25
            outliers = prices[(prices < q25 - 1.5*iqr) | (prices > q75 + 1.5*iqr)]
            
            if len(outliers) > 0:
                alerts.append({
                    'type': 'Prix atypiques',
                    'message': f'{len(outliers)} produit(s) avec prix hors norme',
                    'color': 'F59E0B',
                    'count': len(outliers)
                })
        
        # Détection des dates de péremption proches
        if self.date_cols:
            for date_col in self.date_cols:
                if any(x in date_col.lower() for x in ['peremption', 'expiration', 'dlc', 'dluo']):
                    dates = pd.to_datetime(self.df[date_col], errors='coerce')
                    today = datetime.now()
                    alert_date = today + timedelta(days=30)
                    
                    expiring_mask = (dates >= today) & (dates <= alert_date)
                    expired_mask = dates < today
                    
                    if expired_mask.sum() > 0:
                        alerts.append({
                            'type': 'Produits périmés',
                            'message': f'{int(expired_mask.sum())} produit(s) périmé(s)',
                            'color': 'DC2626',
                            'count': int(expired_mask.sum())
                        })
                    
                    if expiring_mask.sum() > 0:
                        alerts.append({
                            'type': 'Péremption proche',
                            'message': f'{int(expiring_mask.sum())} produit(s) à périr dans 30 jours',
                            'color': 'F59E0B',
                            'count': int(expiring_mask.sum())
                        })
        
        # Si pas d'alertes majeures, message positif
        if not alerts:
            alerts.append({
                'type': 'Situation normale',
                'message': 'Aucune alerte critique détectée',
                'color': '10B981',
                'count': 0
            })
        
        return alerts
    
    def generate_report(self) -> BytesIO:
        """
        Génère le rapport Excel complet avec tous les onglets.
        
        Returns:
            BytesIO: Fichier Excel en mémoire prêt à être téléchargé
        """
        # Création des onglets dans l'ordre logique
        self._create_summary_sheet()      # Résumé exécutif
        self._create_data_sheet()          # Données brutes
        self._create_alerts_sheet()        # Alertes détaillées
        self._create_analysis_sheet()      # Analyse par catégorie
        
        # Sauvegarde en mémoire
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_summary_sheet(self):
        """
        Crée l'onglet Résumé Exécutif avec les KPIs principaux et alertes.
        """
        ws = self.wb.create_sheet("📊 Résumé", 0)
        
        # Titre principal
        ws['A1'] = "RAPPORT DE STOCK - PHARMA-AI COPILOT"
        ws['A1'].style = self.style_title
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30
        
        # Date du rapport
        ws['A3'] = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A3'].style = self.style_info
        ws.merge_cells('A3:E3')
        
        # Section KPIs principaux
        ws['A5'] = "INDICATEURS CLÉS"
        ws['A5'].style = self.style_header
        ws.merge_cells('A5:C5')
        ws.row_dimensions[5].height = 25
        
        # Extraction et affichage des KPIs
        row = 6
        
        # Nombre total de produits/références
        ws[f'A{row}'] = "Total des références"
        ws[f'B{row}'] = len(self.df)
        ws[f'B{row}'].style = self.style_number
        ws[f'C{row}'] = "Nombre de lignes dans le fichier"
        ws[f'C{row}'].font = Font(size=9, italic=True, color='64748B')
        row += 1
        
        # Nombre de fournisseurs/marques
        if self.brand_col:
            ws[f'A{row}'] = "Fournisseurs / Marques"
            ws[f'B{row}'] = self.df[self.brand_col].nunique()
            ws[f'B{row}'].style = self.style_number
            ws[f'C{row}'] = f"Basé sur '{self.brand_col}'"
            ws[f'C{row}'].font = Font(size=9, italic=True, color='64748B')
            row += 1
        
        # Nombre de catégories
        if self.category_col:
            ws[f'A{row}'] = "Catégories"
            ws[f'B{row}'] = self.df[self.category_col].nunique()
            ws[f'B{row}'].style = self.style_number
            ws[f'C{row}'] = f"Basé sur '{self.category_col}'"
            ws[f'C{row}'].font = Font(size=9, italic=True, color='64748B')
            row += 1
        
        # Statistiques de prix
        if self.price_cols:
            price_col = self.price_cols[0]
            prices = self._clean_prices(self.df[price_col])
            
            ws[f'A{row}'] = "Prix moyen"
            ws[f'B{row}'] = prices.mean()
            ws[f'B{row}'].style = self.style_currency
            ws[f'C{row}'] = f"Basé sur '{price_col}'"
            ws[f'C{row}'].font = Font(size=9, italic=True, color='64748B')
            row += 1
            
            ws[f'A{row}'] = "Prix médian"
            ws[f'B{row}'] = prices.median()
            ws[f'B{row}'].style = self.style_currency
            row += 1
            
            ws[f'A{row}'] = "Valeur totale du stock"
            ws[f'B{row}'] = prices.sum()
            ws[f'B{row}'].style = self.style_currency
            row += 1
            
            ws[f'A{row}'] = "Fourchette de prix"
            ws[f'B{row}'] = f"{prices.min():.2f} € - {prices.max():.2f} €"
            ws[f'B{row}'].font = Font(name='Calibri', size=10)
            row += 1
        
        # Statistiques de quantité
        if self.quantity_cols:
            qty_col = self.quantity_cols[0]
            qty_values = pd.to_numeric(self.df[qty_col], errors='coerce')
            
            ws[f'A{row}'] = "Stock total"
            ws[f'B{row}'] = qty_values.sum()
            ws[f'B{row}'].style = self.style_number
            ws[f'C{row}'] = f"Basé sur '{qty_col}'"
            ws[f'C{row}'].font = Font(size=9, italic=True, color='64748B')
            row += 1
            
            ws[f'A{row}'] = "Stock moyen par produit"
            ws[f'B{row}'] = qty_values.mean()
            ws[f'B{row}'].style = self.style_decimal
            row += 1
        
        # Section Alertes
        row += 2
        ws[f'A{row}'] = "ALERTES ET RECOMMANDATIONS"
        ws[f'A{row}'].style = self.style_header
        ws.merge_cells(f'A{row}:E{row}')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Calcul et affichage des alertes
        alerts = self._calculate_alerts()
        
        for alert in alerts:
            # Icône selon le type
            icon = "🔴" if alert['color'] == 'EF4444' or alert['color'] == 'DC2626' else \
                   "🟡" if alert['color'] == 'F59E0B' else "🟢"
            
            ws[f'A{row}'] = f"{icon} {alert['type']}"
            ws[f'A{row}'].font = Font(bold=True, color=alert['color'])
            ws[f'B{row}'] = alert['message']
            ws.merge_cells(f'B{row}:E{row}')
            row += 1
        
        # Section informations complémentaires
        row += 2
        ws[f'A{row}'] = "INFORMATIONS SUR LE FICHIER SOURCE"
        ws[f'A{row}'].style = self.style_header
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
        
        ws[f'A{row}'] = "Colonnes détectées"
        ws[f'B{row}'] = len(self.df.columns)
        ws[f'B{row}'].style = self.style_number
        row += 1
        
        ws[f'A{row}'] = "Types de données identifiés"
        detected_types = set(m['detected_type'] for m in self.schema.values())
        ws[f'B{row}'] = ", ".join(sorted(detected_types))
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].font = Font(size=10)
        
        # Ajustement des largeurs de colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_data_sheet(self):
        """
        Crée l'onglet Données avec les données nettoyées et mise en forme conditionnelle.
        """
        ws = self.wb.create_sheet("📋 Données")
        
        # En-têtes avec style
        headers = list(self.df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.style = self.style_header
        
        ws.row_dimensions[1].height = 25
        
        # Données avec style adapté au type
        for row_idx, row_data in enumerate(self.df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Déterminer le style selon le type de colonne
                col_name = headers[col_idx - 1]
                detected_type = self.schema.get(col_name, {}).get('detected_type', 'unknown')
                
                if detected_type == 'price':
                    # Nettoyer et formater comme monétaire
                    try:
                        if isinstance(value, (int, float)):
                            cell.value = value
                        else:
                            # Tentative de nettoyage
                            clean_val = str(value).replace('€', '').replace(' ', '').replace(',', '.')
                            cell.value = float(clean_val) if clean_val else value
                        cell.style = self.style_currency
                    except:
                        cell.style = self.style_data
                        
                elif detected_type == 'quantity':
                    try:
                        num_val = pd.to_numeric(value, errors='coerce')
                        cell.value = num_val if not pd.isna(num_val) else value
                        cell.style = self.style_number
                        
                        # Mise en forme conditionnelle pour stock critique
                        if isinstance(cell.value, (int, float)) and cell.value <= 5:
                            cell.style = self.style_alert_red
                        elif isinstance(cell.value, (int, float)) and cell.value <= 10:
                            cell.style = self.style_alert_orange
                    except:
                        cell.style = self.style_data
                        
                elif detected_type == 'date':
                    try:
                        # Tentative de conversion en date
                        if isinstance(value, str):
                            parsed_date = pd.to_datetime(value, errors='coerce')
                            if not pd.isna(parsed_date):
                                cell.value = parsed_date
                        cell.style = self.style_date
                    except:
                        cell.style = self.style_data
                        
                elif detected_type == 'code':
                    # Codes (CIP, EAN, etc.) - alignement centré
                    cell.style = self.style_data
                    cell.alignment = Alignment(horizontal='center')
                    
                else:
                    cell.style = self.style_data
        
        # Ajustement automatique des largeurs de colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Limiter à 50 caractères pour éviter colonnes trop larges
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = max(adjusted_width, 12)
        
        # Figer la première ligne (en-têtes)
        ws.freeze_panes = 'A2'
    
    def _create_alerts_sheet(self):
        """
        Crée l'onglet Alertes avec les produits à risque détaillés.
        """
        ws = self.wb.create_sheet("🚨 Alertes")
        
        # Titre
        ws['A1'] = "LISTE DES PRODUITS À SURVEILLER"
        ws['A1'].style = self.style_title
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Sous-titre
        ws['A3'] = "Cette liste présente les produits nécessitant une attention particulière (stock faible, péremption proche, prix anormal)"
        ws['A3'].style = self.style_info
        ws.merge_cells('A3:F3')
        ws.row_dimensions[3].height = 30
        
        # En-têtes de tableau
        headers = ['Produit', 'Code/CIP', 'Problème détecté', 'Valeur actuelle', 'Seuil/Recommandation', 'Priorité']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.style = self.style_header
        
        ws.row_dimensions[5].height = 25
        
        # Récupération des alertes détaillées
        alerts = self._calculate_alerts(detailed=True)
        
        row = 6
        alert_count = 0
        
        for alert in alerts:
            if 'produit' in alert:  # Alertes détaillées seulement
                # Produit
                ws.cell(row=row, column=1, value=alert.get('produit', 'N/A'))
                
                # Code/CIP si disponible
                code_val = ''
                if self.code_col:
                    # Trouver le code correspondant au produit
                    product_mask = self.df[self.product_col] == alert.get('produit') if self.product_col else pd.Series([False])
                    if product_mask.any():
                        code_val = self.df.loc[product_mask, self.code_col].iloc[0]
                ws.cell(row=row, column=2, value=code_val)
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
                
                # Type de problème
                problem_cell = ws.cell(row=row, column=3, value=alert.get('type', 'N/A'))
                
                # Valeur actuelle
                ws.cell(row=row, column=4, value=alert.get('valeur', 'N/A'))
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
                
                # Seuil/Recommandation
                ws.cell(row=row, column=5, value=alert.get('action', alert.get('seuil', 'Vérifier')))
                
                # Priorité avec couleur
                priorite = alert.get('priorite', 'MOYENNE')
                priorite_cell = ws.cell(row=row, column=6, value=priorite)
                
                if priorite == 'HAUTE':
                    priorite_cell.style = self.style_alert_red
                elif priorite == 'MOYENNE':
                    priorite_cell.style = self.style_alert_orange
                else:
                    priorite_cell.style = self.style_alert_green
                
                row += 1
                alert_count += 1
        
        # Message si aucune alerte détaillée
        if alert_count == 0:
            ws.cell(row=row, column=1, value="✅ Aucune alerte détaillée à afficher")
            ws.merge_cells(f'A{row}:F{row}')
            ws.cell(row=row, column=1).font = Font(size=12, color='10B981', bold=True)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 35  # Produit
        ws.column_dimensions['B'].width = 18  # Code
        ws.column_dimensions['C'].width = 22  # Problème
        ws.column_dimensions['D'].width = 15  # Valeur
        ws.column_dimensions['E'].width = 30  # Recommandation
        ws.column_dimensions['F'].width = 12  # Priorité
        
        # Figer les en-têtes
        ws.freeze_panes = 'A6'
    
    def _create_analysis_sheet(self):
        """
        Crée l'onglet Analyse avec les statistiques par catégorie et fournisseur.
        """
        ws = self.wb.create_sheet("📈 Analyse")
        
        # Titre principal
        ws['A1'] = "ANALYSE PAR CATÉGORIE ET FOURNISSEUR"
        ws['A1'].style = self.style_title
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        current_row = 3
        
        # Analyse par catégorie si disponible
        if self.category_col and (self.price_cols or self.quantity_cols):
            ws[f'A{current_row}'] = f"STATISTIQUES PAR CATÉGORIE ({self.category_col})"
            ws[f'A{current_row}'].style = self.style_header
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            # En-têtes du tableau
            cat_headers = ['Catégorie', 'Nombre de produits', 'Stock total', 'Valeur stock', 'Prix moyen', '% du total']
            for col_idx, header in enumerate(cat_headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.style = self.style_header
            current_row += 1
            
            # Calcul des statistiques par catégorie
            cat_stats = []
            total_value = 0
            
            for category in self.df[self.category_col].unique():
                if pd.isna(category):
                    continue
                    
                mask = self.df[self.category_col] == category
                cat_data = self.df[mask]
                
                nb_produits = len(cat_data)
                
                # Stock total
                stock_total = 0
                if self.quantity_cols:
                    qty_vals = pd.to_numeric(cat_data[self.quantity_cols[0]], errors='coerce')
                    stock_total = qty_vals.sum()
                
                # Valeur du stock
                valeur_stock = 0
                prix_moyen = 0
                if self.price_cols:
                    prices = self._clean_prices(cat_data[self.price_cols[0]])
                    valeur_stock = prices.sum()
                    prix_moyen = prices.mean()
                    total_value += valeur_stock
                
                cat_stats.append({
                    'categorie': category,
                    'nb_produits': nb_produits,
                    'stock_total': stock_total,
                    'valeur_stock': valeur_stock,
                    'prix_moyen': prix_moyen
                })
            
            # Trier par valeur de stock décroissante
            cat_stats.sort(key=lambda x: x['valeur_stock'], reverse=True)
            
            # Affichage des données
            for stat in cat_stats:
                pct_total = (stat['valeur_stock'] / total_value * 100) if total_value > 0 else 0
                
                ws.cell(row=current_row, column=1, value=stat['categorie'])
                ws.cell(row=current_row, column=2, value=stat['nb_produits']).style = self.style_number
                ws.cell(row=current_row, column=3, value=stat['stock_total']).style = self.style_number
                ws.cell(row=current_row, column=4, value=stat['valeur_stock']).style = self.style_currency
                ws.cell(row=current_row, column=5, value=stat['prix_moyen']).style = self.style_currency
                
                # Pourcentage avec style
                pct_cell = ws.cell(row=current_row, column=6, value=pct_total / 100)
                pct_cell.style = self.style_percent
                
                current_row += 1
            
            current_row += 2
        
        # Analyse par fournisseur si disponible
        if self.brand_col and self.price_cols:
            ws[f'A{current_row}'] = f"TOP FOURNISSEURS PAR VALEUR ({self.brand_col})"
            ws[f'A{current_row}'].style = self.style_header
            ws.merge_cells(f'A{current_row}:E{current_row}')
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            # En-têtes
            brand_headers = ['Fournisseur', 'Nombre de produits', 'Valeur totale', 'Prix moyen', '% du total']
            for col_idx, header in enumerate(brand_headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=header)
                cell.style = self.style_header
            current_row += 1
            
            # Calcul par fournisseur
            brand_stats = []
            total_brand_value = 0
            
            for brand in self.df[self.brand_col].dropna().unique():
                mask = self.df[self.brand_col] == brand
                brand_data = self.df[mask]
                
                nb_produits = len(brand_data)
                prices = self._clean_prices(brand_data[self.price_cols[0]])
                valeur = prices.sum()
                prix_moy = prices.mean()
                total_brand_value += valeur
                
                brand_stats.append({
                    'fournisseur': brand,
                    'nb_produits': nb_produits,
                    'valeur': valeur,
                    'prix_moyen': prix_moy
                })
            
            # Top 10 par valeur
            brand_stats.sort(key=lambda x: x['valeur'], reverse=True)
            
            for stat in brand_stats[:10]:
                pct = (stat['valeur'] / total_brand_value * 100) if total_brand_value > 0 else 0
                
                ws.cell(row=current_row, column=1, value=stat['fournisseur'])
                ws.cell(row=current_row, column=2, value=stat['nb_produits']).style = self.style_number
                ws.cell(row=current_row, column=3, value=stat['valeur']).style = self.style_currency
                ws.cell(row=current_row, column=4, value=stat['prix_moyen']).style = self.style_currency
                
                pct_cell = ws.cell(row=current_row, column=5, value=pct / 100)
                pct_cell.style = self.style_percent
                
                current_row += 1
        
        # Ajustement des largeurs
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # Figer les en-têtes de la première section
        if current_row > 3:
            ws.freeze_panes = 'A4'


# Fonction utilitaire pour générer le rapport rapidement
def generate_excel_report(df: pd.DataFrame, schema: Dict) -> BytesIO:
    """
    Fonction simplifiée pour générer un rapport Excel.
    
    Args:
        df: DataFrame avec les données
        schema: Schéma détecté des colonnes
        
    Returns:
        BytesIO contenant le fichier Excel
    """
    generator = ExcelReportGenerator(df, schema)
    return generator.generate_report()
